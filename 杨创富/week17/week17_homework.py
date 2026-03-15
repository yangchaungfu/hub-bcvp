import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple


def levenshtein_distance(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev_row = list(range(len(b) + 1))
    for i, ca in enumerate(a, start=1):
        curr_row = [i]
        for j, cb in enumerate(b, start=1):
            ins = curr_row[j - 1] + 1
            delete = prev_row[j] + 1
            replace = prev_row[j - 1] + (0 if ca == cb else 1)
            curr_row.append(min(ins, delete, replace))
        prev_row = curr_row
    return prev_row[-1]


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    dist = levenshtein_distance(a, b)
    return 1.0 - dist / max(len(a), len(b))


def split_values(raw: str) -> List[str]:
    if not raw:
        return []
    parts = re.split(r"[;,|，；、/\s]+", raw.strip())
    return [p for p in parts if p]


def read_excel_rows(excel_path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "读取 Excel 需要 openpyxl。请先执行: pip install openpyxl"
        ) from exc

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in row])
    if not rows:
        return []

    header = rows[0]
    data_rows = rows[1:]
    result: List[Dict[str, str]] = []
    for row in data_rows:
        item = {}
        for i, name in enumerate(header):
            if not name:
                continue
            item[name] = row[i] if i < len(row) else ""
        result.append(item)
    return result


@dataclass
class Node:
    node_id: str
    intents: List[str]
    slots: List[str] = field(default_factory=list)
    actions: List[str] = field(default_factory=list)
    response: str = ""
    child_nodes: List[str] = field(default_factory=list)


@dataclass
class DialogueState:
    scenario_name: str
    current_node_id: Optional[str]
    active_node_id: Optional[str]
    slots: Dict[str, str] = field(default_factory=dict)
    requested_slot: Optional[str] = None
    is_finished: bool = False


class Scenario:
    def __init__(self, path: Path):
        self.path = path
        data = json.loads(path.read_text(encoding="utf-8"))
        self.nodes: Dict[str, Node] = {}
        for item in data:
            node = Node(
                node_id=item["id"],
                intents=item.get("intent", []),
                slots=item.get("slot", []),
                actions=item.get("action", []),
                response=item.get("response", ""),
                child_nodes=item.get("childnode", []),
            )
            self.nodes[node.node_id] = node
        self.root_node_id = data[0]["id"] if data else None

    def get_node(self, node_id: Optional[str]) -> Optional[Node]:
        if not node_id:
            return None
        return self.nodes.get(node_id)


class SlotOntology:
    def __init__(self):
        self.slot_questions: Dict[str, str] = {}
        self.slot_values: Dict[str, List[str]] = {}

    def load_excel(self, excel_path: Path) -> None:
        rows = read_excel_rows(excel_path)
        if not rows:
            return

        def pick_key(keys: List[str], candidates: List[str]) -> Optional[str]:
            for k in keys:
                normalized = k.strip().lower()
                for c in candidates:
                    if c in normalized:
                        return k
            return None

        keys = list(rows[0].keys())
        slot_key = pick_key(keys, ["槽", "slot"])
        ask_key = pick_key(keys, ["反问", "追问", "ask", "question"])
        value_key = pick_key(keys, ["值", "value", "枚举"])
        if not slot_key:
            raise ValueError("Excel 中未找到槽位列，请确保存在 '槽位' 或 'slot' 相关列名。")

        for row in rows:
            slot = row.get(slot_key, "").strip()
            if not slot:
                continue
            slot_name = slot if slot.startswith("#") else f"#{slot}#"
            if ask_key and row.get(ask_key, "").strip():
                self.slot_questions[slot_name] = row.get(ask_key, "").strip()
            if value_key and row.get(value_key, "").strip():
                self.slot_values[slot_name] = split_values(row.get(value_key, ""))

    def get_question(self, slot_name: str) -> str:
        if slot_name in self.slot_questions:
            return self.slot_questions[slot_name]
        pure_name = slot_name.strip("#")
        return f"请告诉我{pure_name}。"


class NLU:
    def __init__(self, ontology: SlotOntology):
        self.ontology = ontology
        self.default_values = {
            "#支付方式#": ["微信", "支付宝", "银行卡", "信用卡", "现金"],
            "#服装颜色#": ["黑", "白", "红", "蓝", "绿", "灰"],
            "#服装尺寸#": ["S", "M", "L", "XL", "XXL"],
            "#服装类型#": ["衬衫", "外套", "裤子", "卫衣", "毛衣", "T恤"],
        }
        # 新增：重听意图关键词
        self.rehear_keywords = ["再说一遍", "重听", "重复", "再说一次", "再说下", "重复一遍"]

    def is_rehear_intent(self, text: str) -> bool:
        """判断用户是否表达重听意图"""
        # 方式1：精准匹配关键词
        for keyword in self.rehear_keywords:
            if keyword in text:
                return True
        # 方式2：相似度匹配（容错，比如用户输入“再说遍”）
        for keyword in self.rehear_keywords:
            if similarity(text, keyword) >= 0.7:
                return True
        return False

    def intent_recognize(self, text: str, candidates: List[Tuple[str, str]]) -> Tuple[Optional[str], float]:
        best_node = None
        best_score = 0.0
        for node_id, intent_text in candidates:
            score = similarity(text, intent_text)
            if intent_text in text:
                score = max(score, 0.95)
            if score > best_score:
                best_score = score
                best_node = node_id
        return best_node, best_score

    def extract_slots(self, text: str, target_slots: List[str]) -> Dict[str, str]:
        result: Dict[str, str] = {}
        for slot in target_slots:
            value = self._extract_one_slot(text, slot)
            if value:
                result[slot] = value
        return result

    def normalize_free_text_slot_value(self, text: str) -> str:
        cleaned = text.strip()
        cleaned = re.sub(r"^[，。！？、,.!?\s]+|[，。！？、,.!?\s]+$", "", cleaned)
        cleaned = re.sub(r"^(我想要|我要|我想买|买|来个|来件|给我来件|给我|要)\s*", "", cleaned)
        return cleaned.strip()

    def _extract_one_slot(self, text: str, slot: str) -> Optional[str]:
        # 优先使用配置枚举值
        candidates = self.ontology.slot_values.get(slot, []) or self.default_values.get(slot, [])
        for candidate in candidates:
            if candidate and candidate in text:
                return candidate
        if candidates:
            token_candidates = re.findall(r"[\u4e00-\u9fa5A-Za-z0-9]+", text)
            best_value = None
            best_score = 0.0
            for token in token_candidates:
                for candidate in candidates:
                    score = similarity(token.lower(), candidate.lower())
                    if score > best_score:
                        best_score = score
                        best_value = candidate
            if best_score >= 0.6:
                return best_value

        # 通用规则
        if slot == "#分期付款期数#":
            m = re.search(r"(\d+)\s*期", text)
            if m:
                return m.group(1)
            m2 = re.search(r"\b(\d+)\b", text)
            if m2:
                return m2.group(1)
        if slot == "#支付方式#":
            for k in ["微信", "支付宝", "银行卡", "信用卡", "现金"]:
                if k in text:
                    return k
        if slot == "#服装尺寸#":
            m = re.search(r"\b(XXL|XL|L|M|S)\b", text, flags=re.IGNORECASE)
            if m:
                return m.group(1).upper()
            m2 = re.search(r"(\d+)\s*码", text)
            if m2:
                return f"{m2.group(1)}码"
        if slot == "#时间#":
            m = re.search(r"(\d{1,2})\s*点", text)
            if m:
                return m.group(1)
        if slot == "#电影名称#":
            m = re.search(r"看(.*?)电影", text)
            if m and m.group(1).strip():
                return m.group(1).strip()
        return None


class DST:
    def update(self, state: DialogueState, recognized_slots: Dict[str, str]) -> DialogueState:
        state.slots.update(recognized_slots)
        return state

    def missing_slots(self, state: DialogueState, node: Node) -> List[str]:
        return [slot for slot in node.slots if slot not in state.slots]


class PM:
    def __init__(self, scenarios: Dict[str, Scenario], nlu: NLU, dst: DST):
        self.scenarios = scenarios
        self.nlu = nlu
        self.dst = dst

    def init_state(self, scenario_name: str) -> DialogueState:
        scenario = self.scenarios[scenario_name]
        return DialogueState(
            scenario_name=scenario_name,
            current_node_id=scenario.root_node_id,
            active_node_id=None,
            slots={},
        )

    def _candidate_nodes(self, state: DialogueState) -> List[str]:
        scenario = self.scenarios[state.scenario_name]
        current = scenario.get_node(state.current_node_id)
        if not current:
            return []
        candidates = [current.node_id]
        candidates.extend(current.child_nodes)
        return candidates

    def step(self, state: DialogueState, user_text: str) -> Tuple[DialogueState, Dict[str, str]]:
        # 新增：优先判断重听意图
        if self.nlu.is_rehear_intent(user_text):
            return state, {"type": "rehear", "text": ""}  # text留空，后续用上一轮回复填充

        scenario = self.scenarios[state.scenario_name]
        if state.is_finished:
            return state, {"type": "end", "text": "流程已经结束，如需继续请重新开始。"}

        candidate_node_ids = self._candidate_nodes(state)
        candidate_intents: List[Tuple[str, str]] = []
        for node_id in candidate_node_ids:
            node = scenario.get_node(node_id)
            if not node:
                continue
            for intent_text in node.intents:
                candidate_intents.append((node_id, intent_text))

        target_node_id, score = self.nlu.intent_recognize(user_text, candidate_intents)
        if target_node_id is None or score < 0.45:
            # 用户可能是在补充槽位信息
            target_node_id = state.active_node_id or state.current_node_id

        node = scenario.get_node(target_node_id)
        if not node:
            return state, {"type": "fallback", "text": "抱歉，我没理解你的意思。"}

        state.active_node_id = node.node_id

        extracted = self.nlu.extract_slots(user_text, node.slots)
        # 当系统正在追问某个槽位时，允许用户以自由文本直接回答，避免反复追问
        if (
            state.requested_slot
            and state.requested_slot in node.slots
            and state.requested_slot not in extracted
        ):
            fallback_value = self.nlu.normalize_free_text_slot_value(user_text)
            if fallback_value:
                extracted[state.requested_slot] = fallback_value
        self.dst.update(state, extracted)
        missing = self.dst.missing_slots(state, node)
        if missing:
            state.requested_slot = missing[0]
            return state, {"type": "ask_slot", "slot": missing[0], "node_id": node.node_id}

        state.requested_slot = None
        state.current_node_id = node.node_id
        if not node.child_nodes:
            state.is_finished = True
        return state, {"type": "respond", "node_id": node.node_id}


class NLG:
    def __init__(self, ontology: SlotOntology):
        self.ontology = ontology

    def render(self, state: DialogueState, scenario: Scenario, policy_output: Dict[str, str], last_response: str = "") -> str:
        """新增last_response参数，用于重听功能"""
        output_type = policy_output.get("type")
        if output_type == "rehear":
            # 重听逻辑：返回上一轮回复，无则提示
            if last_response:
                return f"好的，我再说一遍：{last_response}"
            else:
                return "抱歉，还没有可重复的内容哦。"
        if output_type == "end":
            return policy_output.get("text", "流程结束。")
        if output_type == "fallback":
            return policy_output.get("text", "抱歉，我不太明白。")
        if output_type == "ask_slot":
            slot = policy_output["slot"]
            return self.ontology.get_question(slot)

        node = scenario.get_node(policy_output.get("node_id"))
        if not node:
            return "抱歉，响应生成失败。"

        text = node.response
        for slot_name, slot_value in state.slots.items():
            text = text.replace(slot_name, str(slot_value))
        if node.actions:
            actions_text = "；".join(node.actions)
            return f"{text}\n[执行动作] {actions_text}"
        return text


class DialogueSystem:
    def __init__(self, scenario_dir: Path, excel_path: Optional[Path] = None):
        self.scenario_dir = scenario_dir
        self.scenarios = self._load_scenarios(scenario_dir)
        if not self.scenarios:
            raise RuntimeError(f"未在 {scenario_dir} 中找到场景 JSON。")
        self.ontology = SlotOntology()
        if excel_path and excel_path.exists():
            self.ontology.load_excel(excel_path)
        self.nlu = NLU(self.ontology)
        self.dst = DST()
        self.pm = PM(self.scenarios, self.nlu, self.dst)
        self.nlg = NLG(self.ontology)
        self.state: Optional[DialogueState] = None
        # 新增：记录上一轮系统回复
        self.last_response: str = ""

    def _load_scenarios(self, scenario_dir: Path) -> Dict[str, Scenario]:
        scenarios = {}
        for p in sorted(scenario_dir.glob("*.json")):
            scenario_name = p.stem.replace("scenario-", "")
            scenarios[scenario_name] = Scenario(p)
        return scenarios

    def list_scenarios(self) -> List[str]:
        return list(self.scenarios.keys())

    def start(self, scenario_name: str) -> None:
        if scenario_name not in self.scenarios:
            raise ValueError(f"未知场景: {scenario_name}")
        self.state = self.pm.init_state(scenario_name)
        # 启动场景时清空上一轮回复
        self.last_response = ""

    def chat(self, user_text: str) -> str:
        if not self.state:
            return "请先选择并启动一个场景。"
        state, pm_output = self.pm.step(self.state, user_text)
        self.state = state
        scenario = self.scenarios[self.state.scenario_name]
        # 生成回复时传入上一轮回复
        response = self.nlg.render(self.state, scenario, pm_output, self.last_response)
        # 更新上一轮回复（重听请求不覆盖，避免循环重复）
        if pm_output.get("type") != "rehear":
            self.last_response = response
        return response


def main() -> None:
    base_dir = Path(__file__).parent
    excel_candidates = [
        base_dir / "slot_ontology.xlsx",
        base_dir / "slot_ontology.xls",
    ]
    excel_path = next((p for p in excel_candidates if p.exists()), None)
    ds = DialogueSystem(base_dir, excel_path=excel_path)

    scenarios = ds.list_scenarios()
    print("可用场景：")
    for i, name in enumerate(scenarios, start=1):
        print(f"{i}. {name}")
    choice = input("请输入场景编号: ").strip()
    if not choice.isdigit() or not (1 <= int(choice) <= len(scenarios)):
        print("无效编号，程序结束。")
        return
    scenario_name = scenarios[int(choice) - 1]
    ds.start(scenario_name)

    print(f"已进入场景：{scenario_name}")
    print("输入 quit 退出。支持说「再说一遍/重听」重复上一轮回复。")
    while True:
        user_text = input("你: ").strip()
        if user_text.lower() in {"quit", "exit"}:
            print("系统: 再见。")
            break
        bot = ds.chat(user_text)
        print(f"系统: {bot}")


if __name__ == "__main__":
    main()