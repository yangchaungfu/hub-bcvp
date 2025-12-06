# -*- coding: utf-8 -*-

import json
import re
import os
import torch
import csv  # 导入csv处理模块
import numpy as np
from torch.utils.data import Dataset, DataLoader
from transformers import BertTokenizer

import torch
from openpyxl import load_workbook  # 导入处理xlsx的库
"""
数据加载
"""


class DataGenerator:
    def __init__(self, data_path, config):
        self.config = config
        self.path = data_path
        # self.index_to_label = {0: '家居', 1: '房产', 2: '股票', 3: '社会', 4: '文化',
        #                        5: '国际', 6: '教育', 7: '军事', 8: '彩票', 9: '旅游',
        #                        10: '体育', 11: '科技', 12: '汽车', 13: '健康',
        #                        14: '娱乐', 15: '财经', 16: '时尚', 17: '游戏'}

        self.index_to_label = {0: '0', 1: '1'}

        self.label_to_index = dict((y, x) for x, y in self.index_to_label.items())
        self.config["class_num"] = len(self.index_to_label)
        if self.config["model_type"] == "bert":
            self.tokenizer = BertTokenizer.from_pretrained(config["pretrain_model_path"])
        self.vocab = load_vocab(config["vocab_path"])
        self.config["vocab_size"] = len(self.vocab)

        self.load()


    # def load(self):
    #     self.data = []
    #     with open(self.path, encoding="utf8") as f:
    #         for line in f:
    #             line = json.loads(line)
    #             tag = line["tag"]
    #             label = self.label_to_index[tag]
    #             title = line["title"]
    #             if self.config["model_type"] == "bert":
    #                 input_id = self.tokenizer.encode(title, max_length=self.config["max_length"], pad_to_max_length=True)
    #             else:
    #                 input_id = self.encode_sentence(title)
    #             input_id = torch.LongTensor(input_id)
    #             label_index = torch.LongTensor([label])
    #             self.data.append([input_id, label_index])
    #     return

    def load(self):
        self.data = []
        # 加载xlsx文件，read_only=True提升大文件读取效率
        wb = load_workbook(self.path, read_only=True, data_only=True)
        ws = wb.active  # 获取活动工作表（也可指定表名：wb["Sheet1"]）

        # 遍历行（从第2行开始，跳过表头行）
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 提取第一列（标签）、第二列（review），处理空值避免报错
            tag = str(row[0]) if row[0] is not None else ""
            review = row[1] if row[1] is not None else ""

            # 空数据跳过（可选，避免无效数据）
            if not tag or not review:
                continue

            # 标签转索引（和原逻辑一致）
            label = self.label_to_index[tag]


            # print(label, review)

            # 文本编码（和原JSON逻辑完全一致）
            if self.config["model_type"] == "bert":

                input_id = self.tokenizer.encode(
                    review,
                    max_length=self.config["max_length"],
                    padding= 'max_length',
                    truncation=True  # 截断，避免超长文本报错
                )
            else:
                input_id = self.encode_sentence(review)

            # 转Tensor并添加到数据列表（和原逻辑一致）
            input_id = torch.LongTensor(input_id)
            label_index = torch.LongTensor([label])
            # print(input_id, label_index)
            self.data.append([input_id, label_index])

        # 关闭工作簿释放资源
        wb.close()
        return

    def encode_sentence(self, text):
        input_id = []
        for char in text:
            input_id.append(self.vocab.get(char, self.vocab["[UNK]"]))
        input_id = self.padding(input_id)
        return input_id

    #补齐或截断输入的序列，使其可以在一个batch内运算
    def padding(self, input_id):
        input_id = input_id[:self.config["max_length"]]
        input_id += [0] * (self.config["max_length"] - len(input_id))
        return input_id

    def __len__(self):
        return len(self.data)

    def __getitem__(self, index):
        return self.data[index]

def load_vocab(vocab_path):
    token_dict = {}
    with open(vocab_path, encoding="utf8") as f:
        for index, line in enumerate(f):
            token = line.strip()
            token_dict[token] = index + 1  #0留给padding位置，所以从1开始
    return token_dict


#用torch自带的DataLoader类封装数据
def load_data(data_path, config):
    dg = DataGenerator(data_path, config)
    # print(dg[1])
    # print(dg[2])
    # print(dg[3])
    # print(len(dg.data))
    dl = DataLoader(dg, batch_size=config["batch_size"], shuffle=True)
    return dl

if __name__ == "__main__":
    from config import Config
    dg = DataGenerator("valid_tag_news.json", Config)
    print(dg[1])
